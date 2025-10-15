import os
import datetime
from openpyxl import Workbook, load_workbook
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC



# ============================================================
# 🧱 CREATE OR LOAD EXCEL FILE
# ============================================================
def create_or_load_excel():
    today = datetime.date.today().strftime("%Y-%m-%d")
    
    # Define the reports directory relative to the project root
    # Assuming utils.py is in src/
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    reports_dir = os.path.join(project_root, 'data', 'excel_reports')
    
    # Ensure the directory exists
    os.makedirs(reports_dir, exist_ok=True)
    
    # Construct the full file path
    file_name = os.path.join(reports_dir, f"shipments_{today}.xlsx")
    
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["TrackingNumber", "RecipientName", "Phone", "CommercialValue"])
    return wb, ws, file_name

# ============================================================
# 🧱 HANDLE ALERTS IN BROWSER
# ============================================================
def handle_alert_and_reopen(driver):
    try:
        alert = driver.switch_to.alert
        msg = alert.text
        print(f"⚠️ [alert] Detected: {msg}")
        if "register your user" in msg.lower():
            alert.accept()
            print("🔁 Session expired, redirecting to home...")
            WebDriverWait(driver, 30).until(EC.url_contains("home/applications"))
            return True
        alert.dismiss()
    except NoAlertPresentException:
        pass
    return False



